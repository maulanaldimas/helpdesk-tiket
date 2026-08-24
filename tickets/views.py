import csv
import io
import logging
from datetime import timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.contrib.auth import update_session_auth_hash
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db.models import Avg, Count, ExpressionWrapper, DurationField, F, Q, Sum
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .forms import (
    AppSettingsForm,
    ArticleForm,
    CategoryForm,
    CommentForm,
    CompanyForm,
    ProfileForm,
    RegistrationForm,
    TicketForm,
    UserCreateForm,
    UserEditForm,
)
from .models import Activity, AppSettings, Article, AutoAssignRule, Category, CannedResponse, Comment, Company, InternalNote, Notification, Profile, SatisfactionRating, TimeEntry, Ticket, TicketAttachment
from .tasks import send_notification_email

logger = logging.getLogger('tickets')

MAX_ATTACHMENT_SIZE = 10 * 1024 * 1024  # 10 MB per file


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------

def save_attachments(ticket, files, user):
    """Simpan daftar lampiran ke tiket; abaikan file yang melebihi batas. Return jumlah tersimpan."""
    count = 0
    for f in files or []:
        if not f:
            continue
        if f.size > MAX_ATTACHMENT_SIZE:
            continue
        TicketAttachment.objects.create(ticket=ticket, file=f, uploaded_by=user)
        count += 1
    return count

def notify(user, ticket, message):
    if not user:
        return
    Notification.objects.create(user=user, ticket=ticket, message=message)


def notify_system(users, message):
    """Kirim notifikasi in-app tanpa keterkaitan tiket (mis. persetujuan akun)."""
    for user in users:
        if user:
            notify(user, None, message)


def notify_many(users, ticket, message):
    """Kirim notifikasi in-app ke setiap user, tapi email cuma sekali per alamat unik."""
    sent_emails = set()
    subject = f'[Helpdesk] {message}'
    text_body = (
        f'{message}\n\n'
        f'Tiket: #{ticket.id} - {ticket.title}\n'
        f'Company: {ticket.company.name}\n'
        f'Status: {ticket.get_status_display()}\n\n'
        f'Buka: {settings.SITE_URL}/ticket/{ticket.id}/'
    )
    emails_to_send = []
    for user in users:
        if not user:
            continue
        notify(user, ticket, message)
        if user.email and user.email not in sent_emails:
            emails_to_send.append(user.email)
            sent_emails.add(user.email)

    if not emails_to_send:
        return

    _use_celery = False
    try:
        import redis as _redis_mod
        broker_url = getattr(settings, 'CELERY_BROKER_URL', '')
        if broker_url:
            r = _redis_mod.from_url(broker_url, socket_connect_timeout=2)
            r.ping()
            _use_celery = True
    except Exception:
        pass

    if _use_celery:
        try:
            from .tasks import send_notification_email
            send_notification_email.delay(
                subject=subject,
                text_body=text_body,
                recipient_list=emails_to_send,
                ticket_id=ticket.id,
                message=message,
            )
            return
        except Exception:
            pass

    send_mail(
        subject=subject,
        message=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=emails_to_send,
        fail_silently=True,
    )


def log_activity(ticket, user, action, detail=''):
    Activity.objects.create(ticket=ticket, user=user, action=action, detail=detail[:255])


def get_user_tickets(request):
    """QuerySet tiket yang boleh dilihat user sesuai role-nya."""
    profile = request.user.profile
    if request.user.is_superuser:
        return Ticket.objects.all()
    if profile.role in ('admin', 'staff'):
        return Ticket.objects.filter(company=profile.company)
    return Ticket.objects.filter(created_by=request.user)


def get_visible_ticket(request, pk):
    """Ambil tiket yang boleh diakses user; 404 kalau di luar scope-nya."""
    return get_object_or_404(get_user_tickets(request), pk=pk)


def admin_required(user):
    """True jika superuser atau admin perusahaan."""
    if user.is_superuser:
        return True
    return hasattr(user, 'profile') and user.profile.role == 'admin'


def superuser_required(user):
    """Hanya superuser yang boleh (manajemen lintas perusahaan)."""
    return user.is_superuser


@login_required
def protected_media(request, path):
    """Sajikan lampiran tiket hanya untuk user yang boleh mengakses tiketnya."""
    attachment = get_object_or_404(TicketAttachment, file='tickets/' + path)
    get_visible_ticket(request, attachment.ticket_id)
    return FileResponse(attachment.file.open('rb'))


# ---------------------------------------------------------------------------
# Ticket
# ---------------------------------------------------------------------------

@login_required
def ticket_list(request):
    tickets = get_user_tickets(request)

    query = request.GET.get('q', '').strip()
    if query:
        tickets = tickets.filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )

    tickets = tickets.select_related('company', 'category').order_by('-created_at')

    if request.method == 'POST' and admin_required(request.user):
        action = request.POST.get('bulk_action', '')
        ids = request.POST.getlist('ticket_ids')
        if ids and action:
            selected = Ticket.objects.filter(pk__in=ids)
            if not request.user.is_superuser:
                selected = selected.filter(company=request.user.profile.company)

            if action == 'bulk_status':
                new_status = request.POST.get('bulk_status_value')
                if new_status in dict(Ticket.STATUS_CHOICES):
                    for t in selected:
                        t.status = new_status
                        t.save()
                        log_activity(t, request.user, 'status', f"Bulk: -> {t.get_status_display()}")
                        notify_many({t.created_by, t.assigned_to} - {request.user, None}, t, f"Status tiket #{t.id} diubah menjadi {t.get_status_display()}")
                    messages.success(request, f'{selected.count()} tiket berhasil diubah statusnya.')
            elif action == 'bulk_assign':
                assignee_id = request.POST.get('bulk_assign_value')
                if assignee_id:
                    assignee = User.objects.filter(pk=assignee_id).first()
                    for t in selected:
                        t.assigned_to_id = assignee_id
                        t.save()
                        log_activity(t, request.user, 'assign', f"Bulk: ke {assignee.username}")
                        notify_many({assignee}, t, f"Tiket #{t.id} ditugaskan ke Anda")
                    messages.success(request, f'{selected.count()} tiket berhasil ditugaskan.')
                else:
                    for t in selected:
                        prev = t.assigned_to
                        t.assigned_to = None
                        t.save()
                        log_activity(t, request.user, 'unassign', f"Bulk: dari {prev.username}" if prev else '')
                    messages.success(request, f'{selected.count()} tiket berhasil di-unassign.')
            return redirect('ticket_list')

    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    staff_users = User.objects.none()
    if admin_required(request.user):
        if request.user.is_superuser:
            staff_users = User.objects.filter(profile__role__in=['staff', 'admin'])
        else:
            staff_users = User.objects.filter(profile__role__in=['staff', 'admin'], profile__company=request.user.profile.company)

    return render(request, 'tickets/ticket_list.html', {
        'tickets': page_obj,
        'query': query,
        'staff_users': staff_users,
        'breadcrumbs': [{'label': 'Daftar Tiket'}],
    })


@login_required
def ticket_detail(request, pk):
    ticket = get_visible_ticket(request, pk)
    profile = request.user.profile
    can_manage = profile.role in ['admin', 'staff']

    if request.method == 'POST':
        if 'rate_submit' in request.POST and not can_manage:
            rating_val = request.POST.get('rating')
            if rating_val and ticket.status in ('resolved', 'closed') and ticket.created_by == request.user:
                if not SatisfactionRating.objects.filter(ticket=ticket).exists():
                    SatisfactionRating.objects.create(
                        ticket=ticket,
                        rating=int(rating_val),
                        comment=request.POST.get('rating_comment', ''),
                        created_by=request.user,
                    )
                    log_activity(ticket, request.user, 'comment', f"Rating kepuasan: {rating_val}/5")
                    messages.success(request, 'Terima kasih atas rating Anda!')
            return redirect('ticket_detail', pk=ticket.pk)

        if 'comment_submit' in request.POST:
            comment_form = CommentForm(request.POST, request.FILES)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.ticket = ticket
                comment.author = request.user
                comment.save()
                log_activity(ticket, request.user, 'comment', comment.message[:255])

                if not ticket.first_response_at and request.user.profile.role in ('staff', 'admin'):
                    ticket.first_response_at = timezone.now()
                    ticket.save(update_fields=['first_response_at'])

                added = save_attachments(ticket, comment_form.cleaned_data.get('files') or [], request.user)
                if added:
                    log_activity(ticket, request.user, 'attachment', f"{added} file")

                recipients = {ticket.created_by, ticket.assigned_to}
                recipients.discard(request.user)
                recipients.discard(None)
                notify_many(recipients, ticket, f"{request.user.username} menambahkan komentar di tiket #{ticket.id}")

                return redirect('ticket_detail', pk=ticket.pk)

        elif 'internal_note_submit' in request.POST and can_manage:
            note_msg = request.POST.get('internal_note', '').strip()
            if note_msg:
                InternalNote.objects.create(ticket=ticket, author=request.user, message=note_msg)
                log_activity(ticket, request.user, 'comment', f"[Catatan Internal] {note_msg[:200]}")
                messages.success(request, 'Catatan internal ditambahkan.')
            return redirect('ticket_detail', pk=ticket.pk)

        elif 'status_submit' in request.POST and can_manage:
            new_status = request.POST.get('status')
            valid_statuses = dict(Ticket.STATUS_CHOICES)
            if new_status in valid_statuses and new_status != ticket.status:
                old_display = ticket.get_status_display()
                ticket.status = new_status
                ticket.save()
                log_activity(
                    ticket, request.user, 'status',
                    f"{old_display} -> {ticket.get_status_display()}",
                )
                notify_many({ticket.created_by}, ticket, f"Status tiket #{ticket.id} diubah menjadi {ticket.get_status_display()}")
                return redirect('ticket_detail', pk=ticket.pk)

        elif 'assign_submit' in request.POST and can_manage:
            assignee_id = request.POST.get('assigned_to')
            previous = ticket.assigned_to
            if assignee_id:
                ticket.assigned_to_id = assignee_id
                action = 'assign'
            else:
                ticket.assigned_to = None
                action = 'unassign'
            ticket.save()
            if ticket.assigned_to:
                log_activity(ticket, request.user, action, f"ke {ticket.assigned_to.username}")
                notify_many({ticket.assigned_to}, ticket, f"Kamu ditugaskan ke tiket #{ticket.id}")
            else:
                log_activity(ticket, request.user, action, f"dari {previous.username}" if previous else '')
            if action == 'unassign' and previous:
                notify_many({previous}, ticket, f"Anda tidak lagi menangani tiket #{ticket.id}")
            return redirect('ticket_detail', pk=ticket.pk)

        elif 'sla_pause_submit' in request.POST and can_manage:
            if ticket.sla_paused:
                if ticket.sla_paused_at:
                    paused_seconds = (timezone.now() - ticket.sla_paused_at).total_seconds()
                    ticket.sla_total_paused_seconds += int(paused_seconds)
                ticket.sla_paused = False
                ticket.sla_paused_at = None
                ticket.sla_pause_reason = ''
                log_activity(ticket, request.user, 'status', 'SLA diaktifkan kembali')
            else:
                ticket.sla_paused = True
                ticket.sla_paused_at = timezone.now()
                ticket.sla_pause_reason = request.POST.get('sla_pause_reason', 'Menunggu balasan')
                log_activity(ticket, request.user, 'status', f"SLA dijeda: {ticket.sla_pause_reason}")
            ticket.save()
            return redirect('ticket_detail', pk=ticket.pk)

        elif 'time_start_submit' in request.POST and can_manage:
            active = TimeEntry.objects.filter(ticket=ticket, user=request.user, stopped_at__isnull=True).first()
            if not active:
                TimeEntry.objects.create(
                    ticket=ticket,
                    user=request.user,
                    description=request.POST.get('time_description', ''),
                )
                messages.success(request, 'Timer dimulai.')
            return redirect('ticket_detail', pk=ticket.pk)

        elif 'time_stop_submit' in request.POST and can_manage:
            active = TimeEntry.objects.filter(ticket=ticket, user=request.user, stopped_at__isnull=True).first()
            if active:
                active.stop()
                messages.success(request, f'Timer dihentikan: {active.duration_minutes} menit.')
            return redirect('ticket_detail', pk=ticket.pk)

    comment_form = CommentForm()
    comments = ticket.comments.all().order_by('created_at')
    activities = ticket.activities.select_related('user').order_by('-created_at')[:30]
    attachments = ticket.attachments.order_by('-uploaded_at')
    internal_notes = ticket.internal_notes.select_related('author').order_by('-created_at') if can_manage else []
    time_entries = ticket.time_entries.select_related('user').order_by('-started_at')[:20]
    active_timer = TimeEntry.objects.filter(ticket=ticket, user=request.user, stopped_at__isnull=True).first() if can_manage else None
    total_minutes = ticket.time_entries.aggregate(total=Sum('duration_minutes'))['total'] or 0

    satisfaction = getattr(ticket, 'satisfaction_rating', None)
    can_rate = (
        not can_manage
        and ticket.status in ('resolved', 'closed')
        and ticket.created_by == request.user
        and not satisfaction
    )

    staff_users = []
    canned_responses = []
    if can_manage:
        staff_users = User.objects.filter(
            profile__role__in=['staff', 'admin'],
            profile__company=ticket.company,
        )
        canned_responses = CannedResponse.objects.filter(
            Q(company=ticket.company) | Q(company__isnull=True)
        ).select_related('category')

    return render(request, 'tickets/ticket_detail.html', {
        'ticket': ticket,
        'comments': comments,
        'activities': activities,
        'attachments': attachments,
        'comment_form': comment_form,
        'can_manage': can_manage,
        'staff_users': staff_users,
        'canned_responses': canned_responses,
        'internal_notes': internal_notes,
        'time_entries': time_entries,
        'active_timer': active_timer,
        'total_minutes': total_minutes,
        'satisfaction': satisfaction,
        'can_rate': can_rate,
        'breadcrumbs': [
            {'label': 'Tiket', 'url': '/'},
            {'label': f'#{ticket.id}'},
        ],
    })


@login_required
def ticket_create(request):
    profile = request.user.profile
    if request.user.is_superuser:
        locked_company = None
    else:
        locked_company = profile.company

    if request.method == 'POST':
        form = TicketForm(request.POST, request.FILES)
        if locked_company:
            form.fields['company'].queryset = Company.objects.filter(id=locked_company.id)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            if locked_company:
                ticket.company = locked_company
            ticket.save()
            log_activity(ticket, request.user, 'created', f"Prioritas {ticket.get_priority_display()}")
            added = save_attachments(ticket, form.cleaned_data.get('files') or [], request.user)
            if added:
                log_activity(ticket, request.user, 'attachment', f"{added} file")

            if not ticket.assigned_to:
                auto_user = AutoAssignRule.find_match(ticket)
                if auto_user:
                    ticket.assigned_to = auto_user
                    ticket.save()
                    log_activity(ticket, None, 'assign', f"Otomatis ke {auto_user.username}")
                    notify_many({auto_user}, ticket, f"Tiket #{ticket.id} ditugaskan otomatis ke Anda")

            company_admins = User.objects.filter(
                profile__company=ticket.company, profile__role__in=['staff', 'admin']
            ).exclude(pk=request.user.pk)
            if company_admins:
                notify_many(set(company_admins), ticket, f"Tiket baru #{ticket.id}: {ticket.title}")

            logger.info("Ticket #%s dibuat oleh %s", ticket.id, request.user.username)
            return redirect('ticket_detail', pk=ticket.pk)
    else:
        initial = {}
        if locked_company:
            initial['company'] = locked_company
        form = TicketForm(initial=initial)
        if locked_company:
            form.fields['company'].queryset = Company.objects.filter(id=locked_company.id)
    return render(request, 'tickets/ticket_form.html', {
        'form': form,
        'locked_company': locked_company,
    })


@login_required
def ticket_merge(request, pk):
    """Gabungkan tiket ini ke tiket lain. Semua komentar, lampiran, aktivitas dipindah."""
    ticket = get_visible_ticket(request, pk)
    if not admin_required(request.user):
        return redirect('ticket_detail', pk=pk)

    company_tickets = get_user_tickets(request).exclude(pk=pk).order_by('-created_at')

    if request.method == 'POST':
        target_id = request.POST.get('target_ticket')
        if not target_id:
            messages.error(request, 'Pilih tiket target untuk digabungkan.')
            return redirect('ticket_merge', pk=pk)
        try:
            target = company_tickets.get(pk=target_id)
        except Ticket.DoesNotExist:
            messages.error(request, 'Tiket target tidak ditemukan atau tidak bisa diakses.')
            return redirect('ticket_merge', pk=pk)

        if target.company_id != ticket.company_id:
            messages.error(request, 'Tiket harus dari perusahaan yang sama.')
            return redirect('ticket_merge', pk=pk)

        merge_note = f"[Digabungkan dari tiket #{ticket.id}]"

        Comment.objects.filter(ticket=ticket).update(ticket=target)
        TicketAttachment.objects.filter(ticket=ticket).update(ticket=target)
        Activity.objects.filter(ticket=ticket).update(ticket=target)

        Ticket.objects.filter(pk=ticket.pk).update(
            status='closed',
            closed_at=timezone.now(),
            description=F('description') + f'\n\n{merge_note}',
        )

        log_activity(target, request.user, 'comment', merge_note)
        messages.success(request, f'Tiket #{ticket.id} berhasil digabungkan ke #{target.id}. Tiket #{ticket.id} ditutup.')
        return redirect('ticket_detail', pk=target.pk)

    return render(request, 'tickets/ticket_merge.html', {
        'ticket': ticket,
        'candidates': company_tickets[:50],
    })


@login_required
def import_template(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="template-import-tiket.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['title', 'description', 'company', 'category', 'priority', 'status'])
    writer.writerow(['Laptop tidak menyala', 'Deskripsi masalah lengkap di sini...',
                     'Acme Corp', 'Hardware', 'high', 'open'])
    writer.writerow(['Printer bermasalah', '', 'Acme Corp', 'Hardware', 'medium', 'in_progress'])
    return response


@login_required
def import_tickets(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    if request.method == 'POST':
        uploaded = request.FILES.get('file')
        if not uploaded:
            messages.error(request, 'Pilih file CSV terlebih dahulu.')
            return redirect('import_tickets')
        if uploaded.size > 2 * 1024 * 1024:
            messages.error(request, 'File terlalu besar (maks 2 MB).')
            return redirect('import_tickets')
        try:
            decoded = uploaded.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            messages.error(request, 'File harus berformat CSV ber-encoding UTF-8.')
            return redirect('import_tickets')

        reader = csv.DictReader(io.StringIO(decoded))
        valid_priorities = dict(Ticket.PRIORITY_CHOICES)
        valid_statuses = dict(Ticket.STATUS_CHOICES)
        created = 0
        skipped = []

        for idx, row in enumerate(reader, start=2):
            title = (row.get('title') or '').strip()
            company_name = (row.get('company') or '').strip()
            if not title or not company_name:
                skipped.append(f'baris {idx}: judul/company kosong')
                continue
            if request.user.is_superuser:
                company, _ = Company.objects.get_or_create(name=company_name)
            else:
                company = request.user.profile.company
            cat_name = (row.get('category') or '').strip()
            category = None
            if cat_name:
                category = Category.objects.filter(name__iexact=cat_name).first()
                if not category:
                    category = Category.objects.create(name=cat_name)
            priority = (row.get('priority') or 'medium').strip().lower()
            priority = priority if priority in valid_priorities else 'medium'
            status = (row.get('status') or 'open').strip().lower()
            status = status if status in valid_statuses else 'open'
            Ticket.objects.create(
                title=title,
                description=(row.get('description') or '').strip() or '-',
                company=company,
                category=category,
                priority=priority,
                status=status,
                created_by=request.user,
            )
            created += 1

        messages.success(request, f'{created} tiket berhasil diimpor.')
        if skipped:
            messages.error(request, 'Dilewati: ' + '; '.join(skipped[:10]))
        return redirect('ticket_list')

    return render(request, 'tickets/import_tickets.html')


@login_required
def my_dashboard(request):
    """Ringkasan personal: staff melihat tiket yang ditugaskan, requester tiket buatannya."""
    profile = request.user.profile
    if profile.role == 'staff':
        tickets = Ticket.objects.filter(assigned_to=request.user, company=profile.company)
        scope_label = 'tiket yang ditugaskan ke saya'
    else:
        tickets = Ticket.objects.filter(created_by=request.user)
        scope_label = 'tiket yang saya buat'

    total = tickets.count()
    open_count = tickets.filter(status__in=['open', 'in_progress']).count()
    overdue_count = tickets.exclude(status__in=['resolved', 'closed']).filter(
        sla_deadline__lt=timezone.now()
    ).count()
    resolved_count = tickets.filter(status__in=['resolved', 'closed']).count()

    recent = tickets.select_related('company', 'category').order_by('-updated_at')[:5]

    return render(request, 'tickets/my_dashboard.html', {
        'scope_label': scope_label,
        'total': total,
        'open_count': open_count,
        'overdue_count': overdue_count,
        'resolved_count': resolved_count,
        'recent': recent,
        'breadcrumbs': [{'label': 'Ringkasan Saya'}],
    })


@login_required
def notification_list(request):
    notifications = request.user.notifications.all().order_by('-created_at')
    notifications.filter(is_read=False).update(is_read=True)

    paginator = Paginator(notifications, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'tickets/notification_list.html', {'notifications': page_obj})


# ---------------------------------------------------------------------------
# Laporan
# ---------------------------------------------------------------------------

def get_filtered_tickets(request):
    """Ambil tiket sesuai role, lalu terapkan filter dari query string."""
    tickets = get_user_tickets(request)

    status = request.GET.get('status')
    priority = request.GET.get('priority')
    company_id = request.GET.get('company')

    if status:
        tickets = tickets.filter(status=status)
    if priority:
        tickets = tickets.filter(priority=priority)
    if company_id:
        tickets = tickets.filter(company_id=company_id)

    return tickets.select_related('company', 'category', 'created_by', 'assigned_to').order_by('-created_at')


@login_required
def report_page(request):
    tickets = get_filtered_tickets(request)
    profile = request.user.profile

    companies = Company.objects.filter(id=profile.company_id) if not request.user.is_superuser else Company.objects.all()

    return render(request, 'tickets/report.html', {
        'tickets': tickets,
        'companies': companies,
        'status_choices': Ticket.STATUS_CHOICES,
        'priority_choices': Ticket.PRIORITY_CHOICES,
    })


@login_required
def report_export_excel(request):
    tickets = get_filtered_tickets(request)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Tiket"

    headers = ['ID', 'Judul', 'Company', 'Kategori', 'Status', 'Prioritas', 'Dibuat Oleh', 'Assigned To', 'Dibuat Pada', 'SLA Deadline', 'Overdue']
    ws.append(headers)

    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for t in tickets:
        ws.append([
            t.id,
            t.title,
            t.company.name,
            t.category.name if t.category else '-',
            t.get_status_display(),
            t.get_priority_display(),
            t.created_by.username,
            t.assigned_to.username if t.assigned_to else '-',
            t.created_at.strftime('%Y-%m-%d %H:%M'),
            t.sla_deadline.strftime('%Y-%m-%d %H:%M') if t.sla_deadline else '-',
            'Ya' if t.is_overdue() else 'Tidak',
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"laporan-tiket-{timezone.now().strftime('%Y%m%d-%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def report_export_csv(request):
    tickets = get_filtered_tickets(request)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"laporan-tiket-{timezone.now().strftime('%Y%m%d-%H%M')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # BOM agar Excel membaca UTF-8 dengan benar

    writer = csv.writer(response)
    writer.writerow(['ID', 'Judul', 'Company', 'Kategori', 'Status', 'Prioritas',
                     'Dibuat Oleh', 'Assigned To', 'Dibuat Pada', 'SLA Deadline', 'Overdue'])
    for t in tickets:
        writer.writerow([
            t.id,
            t.title,
            t.company.name,
            t.category.name if t.category else '-',
            t.get_status_display(),
            t.get_priority_display(),
            t.created_by.username,
            t.assigned_to.username if t.assigned_to else '-',
            t.created_at.strftime('%Y-%m-%d %H:%M'),
            t.sla_deadline.strftime('%Y-%m-%d %H:%M') if t.sla_deadline else '-',
            'Ya' if t.is_overdue() else 'Tidak',
        ])
    return response


@login_required
def report_export_pdf(request):
    tickets = get_filtered_tickets(request)

    response = HttpResponse(content_type='application/pdf')
    filename = f"laporan-tiket-{timezone.now().strftime('%Y%m%d-%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Laporan Tiket Helpdesk", styles['Title'])]
    elements.append(Paragraph(f"Dicetak: {timezone.now().strftime('%d %B %Y %H:%M')}", styles['Normal']))
    elements.append(Paragraph(" ", styles['Normal']))

    data = [['ID', 'Judul', 'Company', 'Status', 'Prioritas', 'Assigned To', 'Dibuat Pada', 'Overdue']]
    for t in tickets:
        data.append([
            str(t.id),
            t.title[:30],
            t.company.name,
            t.get_status_display(),
            t.get_priority_display(),
            t.assigned_to.username if t.assigned_to else '-',
            t.created_at.strftime('%d/%m/%Y'),
            'Ya' if t.is_overdue() else 'Tidak',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    return response


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@login_required
def dashboard(request):
    profile = request.user.profile
    tickets = get_user_tickets(request)

    total = tickets.count()

    status_counts = {value: tickets.filter(status=value).count() for value, label in Ticket.STATUS_CHOICES}
    priority_counts = {value: tickets.filter(priority=value).count() for value, label in Ticket.PRIORITY_CHOICES}

    status_data = [
        {'value': value, 'label': label, 'count': status_counts.get(value, 0)}
        for value, label in Ticket.STATUS_CHOICES
    ]
    priority_data = [
        {'value': value, 'label': label, 'count': priority_counts.get(value, 0)}
        for value, label in Ticket.PRIORITY_CHOICES
    ]

    overdue_count = tickets.exclude(status__in=['resolved', 'closed']).filter(
        sla_deadline__lt=timezone.now()
    ).count()

    category_breakdown = (
        tickets.values('category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )

    # SLA compliance & rata-rata waktu penyelesaian
    closed_qs = tickets.filter(status__in=['resolved', 'closed'])
    total_closed = closed_qs.count()
    on_time_count = sum(
        1 for t in closed_qs.only('closed_at', 'sla_deadline')
        if t.closed_at and t.sla_deadline and t.closed_at <= t.sla_deadline
    )
    sla_compliance = round(on_time_count / total_closed * 100) if total_closed else 0

    avg_duration = closed_qs.filter(closed_at__isnull=False).annotate(
        dur=ExpressionWrapper(F('closed_at') - F('created_at'), output_field=DurationField())
    ).aggregate(avg=Avg('dur'))['avg']
    avg_resolution_hours = round(avg_duration.total_seconds() / 3600, 1) if avg_duration else 0

    frt_qs = tickets.filter(first_response_at__isnull=False).annotate(
        frt=ExpressionWrapper(F('first_response_at') - F('created_at'), output_field=DurationField())
    )
    avg_frt = frt_qs.aggregate(avg=Avg('frt'))['avg']
    avg_first_response_hours = round(avg_frt.total_seconds() / 3600, 1) if avg_frt else 0
    total_with_frt = frt_qs.count()

    csat_qs = SatisfactionRating.objects.filter(ticket__in=tickets)
    total_ratings = csat_qs.count()
    avg_csat = round(csat_qs.aggregate(avg=Avg('rating'))['avg'], 1) if total_ratings else 0

    # Tren 14 hari terakhir
    trend_labels, trend_counts = [], []
    today = timezone.localdate()
    for i in range(13, -1, -1):
        day = today - timedelta(days=i)
        trend_labels.append(day.strftime('%d/%m'))
        trend_counts.append(tickets.filter(created_at__date=day).count())

    # Beban per staff (admin)
    staff_workload = []
    if profile.role == 'admin':
        staff_workload = list(
            Ticket.objects.filter(assigned_to__isnull=False, company=profile.company)
            .values('assigned_to__username', 'assigned_to_id')
            .annotate(total=Count('id'))
            .order_by('-total')
        )

    recent_tickets = tickets.select_related('company', 'category').order_by('-created_at')[:5]

    return render(request, 'tickets/dashboard.html', {
        'total': total,
        'status_counts': status_counts,
        'priority_counts': priority_counts,
        'status_data': status_data,
        'priority_data': priority_data,
        'overdue_count': overdue_count,
        'category_breakdown': category_breakdown,
        'sla_compliance': sla_compliance,
        'avg_resolution_hours': avg_resolution_hours,
        'avg_first_response_hours': avg_first_response_hours,
        'total_with_frt': total_with_frt,
        'avg_csat': avg_csat,
        'total_ratings': total_ratings,
        'trend_labels': trend_labels,
        'trend_counts': trend_counts,
        'staff_workload': staff_workload,
        'recent_tickets': recent_tickets,
        'breadcrumbs': [{'label': 'Dashboard'}],
    })


# ---------------------------------------------------------------------------
# Master: Company & Category
# ---------------------------------------------------------------------------

@login_required
def company_list(request):
    if not superuser_required(request.user):
        return redirect('dashboard')

    if request.method == 'POST':
        form = CompanyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company berhasil ditambahkan.')
            return redirect('company_list')
    else:
        form = CompanyForm()

    companies = Company.objects.all().order_by('name')
    return render(request, 'tickets/company_list.html', {'companies': companies, 'form': form})


@login_required
def company_delete(request, pk):
    if not superuser_required(request.user):
        return redirect('dashboard')

    company = get_object_or_404(Company, pk=pk)
    if request.method == 'POST':
        company.delete()
        messages.success(request, 'Company berhasil dihapus.')
    return redirect('company_list')


@login_required
def category_list(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Kategori berhasil ditambahkan.')
            return redirect('category_list')
    else:
        form = CategoryForm()

    categories = Category.objects.all().order_by('name')
    return render(request, 'tickets/category_list.html', {'categories': categories, 'form': form})


@login_required
def category_delete(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Kategori berhasil dihapus.')
    return redirect('category_list')


# ---------------------------------------------------------------------------
# Manajemen User & Role
# ---------------------------------------------------------------------------

@login_required
def user_list(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    role_filter = request.GET.get('role', '')
    users = User.objects.select_related('profile', 'profile__company').all()
    if not request.user.is_superuser:
        users = users.filter(profile__company=request.user.profile.company)
    if role_filter:
        users = users.filter(profile__role=role_filter)

    return render(request, 'tickets/user_list.html', {
        'users': users,
        'role_filter': role_filter,
        'role_choices': [('', 'Semua')] + Profile.ROLE_CHOICES,
    })


@login_required
def user_create(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    locked_company = request.user.profile.company if not request.user.is_superuser else None

    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if locked_company:
            form.fields['company'].queryset = Company.objects.filter(id=locked_company.id)
        if form.is_valid():
            user = form.save()
            if locked_company:
                Profile.objects.filter(user=user).update(company=locked_company)
            messages.success(request, f"User '{form.cleaned_data['username']}' berhasil dibuat.")
            return redirect('user_list')
    else:
        form = UserCreateForm()
        if locked_company:
            form.fields['company'].queryset = Company.objects.filter(id=locked_company.id)

    return render(request, 'tickets/user_form.html', {'form': form, 'title': 'Tambah User'})


@login_required
def user_edit(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    user = get_object_or_404(User, pk=pk)
    if not request.user.is_superuser and user.profile.company != request.user.profile.company:
        return redirect('user_list')
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, f"User '{user.username}' berhasil diupdate.")
            return redirect('user_list')
    else:
        form = UserEditForm(instance=user)

    return render(request, 'tickets/user_form.html', {'form': form, 'title': f'Edit User: {user.username}'})


@login_required
def user_delete(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    user = get_object_or_404(User, pk=pk)
    if not request.user.is_superuser and user.profile.company != request.user.profile.company:
        return redirect('user_list')
    if user == request.user:
        messages.error(request, 'Tidak bisa menghapus akun sendiri.')
        return redirect('user_list')

    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f"User '{username}' berhasil dihapus.")
    return redirect('user_list')


def register(request):
    """Registrasi mandiri: akun dibuat nonaktif sampai disetujui admin."""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if not settings.REGISTRATION_OPEN:
        return render(request, 'tickets/register.html', {'registration_closed': True})

    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            user = User.objects.create_user(
                username=data['username'],
                email=data['email'],
                password=data['password1'],
                first_name=data['first_name'],
                is_active=False,
            )
            Profile.objects.update_or_create(user=user, defaults={
                'company': data['company'],
                'role': 'requester',
                'pending_approval': True,
            })
            admins = User.objects.filter(is_superuser=True) | User.objects.filter(
                profile__role='admin', profile__company=data['company']
            )
            notify_system(admins.distinct(), f"Registrasi baru menunggu persetujuan: {user.username}")
            logger.info("Registrasi baru: %s (menunggu persetujuan)", user.username)
            return render(request, 'tickets/register.html', {'registered': True})
    else:
        form = RegistrationForm()

    return render(request, 'tickets/register.html', {'form': form})


@login_required
def pending_approvals(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    pending = User.objects.filter(profile__pending_approval=True).select_related('profile', 'profile__company')
    if not request.user.is_superuser:
        pending = pending.filter(profile__company=request.user.profile.company)
    return render(request, 'tickets/pending_approvals.html', {'pending': pending})


@login_required
def approve_user(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    user = get_object_or_404(User, pk=pk)
    if not request.user.is_superuser and hasattr(user, 'profile') and user.profile.company != request.user.profile.company:
        return redirect('pending_approvals')
    if request.method == 'POST':
        profile, _ = Profile.objects.get_or_create(user=user)
        profile.pending_approval = False
        profile.save()
        user.profile = profile
        user.is_active = True
        user.save()
        notify_system([user], f"Akun kamu ({user.username}) telah disetujui. Silakan masuk.")
        messages.success(request, f"Akun '{user.username}' disetujui dan aktif.")
    return redirect('pending_approvals')


@login_required
def reject_user(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    user = get_object_or_404(User, pk=pk)
    if not request.user.is_superuser and hasattr(user, 'profile') and user.profile.company != request.user.profile.company:
        return redirect('pending_approvals')
    if request.method == 'POST':
        username = user.username
        user.delete()
        messages.success(request, f"Permintaan '{username}' ditolak dan dihapus.")
    return redirect('pending_approvals')


@login_required
def reset_password(request, pk):
    """Admin mereset password user ke nilai sementara."""
    if not admin_required(request.user):
        return redirect('dashboard')

    user = get_object_or_404(User, pk=pk)
    if not request.user.is_superuser and hasattr(user, 'profile') and user.profile.company != request.user.profile.company:
        return redirect('user_list')
    if request.method == 'POST':
        new_password = request.POST.get('new_password', '').strip()
        if len(new_password) < 8:
            messages.error(request, 'Password minimal 8 karakter.')
        else:
            user.set_password(new_password)
            user.save()
            messages.success(request, f"Password untuk '{user.username}' berhasil direset.")
            return redirect('user_list')
    return render(request, 'tickets/reset_password.html', {'target_user': user})


@login_required
def settings_page(request):
    if not superuser_required(request.user):
        return redirect('dashboard')

    cfg = AppSettings.load()
    if request.method == 'POST':
        form = AppSettingsForm(request.POST, request.FILES, instance=cfg)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pengaturan aplikasi disimpan.')
            return redirect('settings_page')
    else:
        form = AppSettingsForm(instance=cfg)

    return render(request, 'tickets/settings_form.html', {'form': form})


@login_required
def profile_page(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, 'Profil berhasil diperbarui.')
            return redirect('profile')
    else:
        form = ProfileForm(instance=profile)

    return render(request, 'tickets/profile_form.html', {'form': form, 'profile': profile})


@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, 'Password berhasil diubah.')
            return redirect('dashboard')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'tickets/change_password.html', {'form': form})


# ---------------------------------------------------------------------------
# Knowledge Base / FAQ
# ---------------------------------------------------------------------------

def article_list(request):
    articles = Article.objects.filter(is_published=True)

    query = request.GET.get('q', '').strip()
    category_id = request.GET.get('category', '')
    if query:
        articles = articles.filter(
            Q(title__icontains=query) | Q(content__icontains=query)
        )
    if category_id:
        articles = articles.filter(category_id=category_id)

    paginator = Paginator(articles.select_related('category', 'author'), 12)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'tickets/article_list.html', {
        'articles': page_obj,
        'query': query,
        'categories': Category.objects.all(),
        'selected_category': category_id,
    })


def article_detail(request, pk):
    article = get_object_or_404(
        Article.objects.select_related('author', 'category'),
        pk=pk,
        is_published=True,
    )
    related = (
        Article.objects.filter(is_published=True)
        .exclude(pk=article.pk)
        .order_by('-updated_at')[:5]
    )
    return render(request, 'tickets/article_detail.html', {
        'article': article,
        'related': related,
    })


@login_required
@require_POST
def article_preview(request):
    """Render Markdown ke HTML untuk pratinjau di editor artikel."""
    if not admin_required(request.user):
        return redirect('dashboard')
    html = Article(content=request.POST.get('content', '')).content_html()
    return HttpResponse(html)


@login_required
def article_create(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    if request.method == 'POST':
        form = ArticleForm(request.POST)
        if form.is_valid():
            article = form.save(commit=False)
            article.author = request.user
            article.save()
            messages.success(request, 'Artikel berhasil dipublikasikan.')
            return redirect('article_detail', pk=article.pk)
    else:
        form = ArticleForm()

    return render(request, 'tickets/article_form.html', {'form': form, 'title': 'Tulis Artikel'})


@login_required
def article_edit(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    article = get_object_or_404(Article, pk=pk)
    if request.method == 'POST':
        form = ArticleForm(request.POST, instance=article)
        if form.is_valid():
            form.save()
            messages.success(request, 'Artikel berhasil diupdate.')
            return redirect('article_detail', pk=article.pk)
    else:
        form = ArticleForm(instance=article)

    return render(request, 'tickets/article_form.html', {'form': form, 'title': f'Edit Artikel: {article.title}'})


@login_required
def article_delete(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    article = get_object_or_404(Article, pk=pk)
    if request.method == 'POST':
        article.delete()
        messages.success(request, 'Artikel berhasil dihapus.')
    return redirect('article_list')


# ---------------------------------------------------------------------------
# Activity Log (audit trail)
# ---------------------------------------------------------------------------

@login_required
def activity_log(request):
    """Halaman audit trail terpusat — superuser lihat semua, admin/staff lihat company sendiri."""
    if not admin_required(request.user):
        return redirect('dashboard')

    activities = Activity.objects.select_related('ticket', 'user', 'ticket__company').order_by('-created_at')
    if not request.user.is_superuser:
        activities = activities.filter(ticket__company=request.user.profile.company)

    action_filter = request.GET.get('action', '')
    if action_filter:
        activities = activities.filter(action=action_filter)

    paginator = Paginator(activities, 30)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'tickets/activity_log.html', {
        'activities': page_obj,
        'action_filter': action_filter,
        'action_choices': Activity.ACTION_CHOICES,
    })


# ---------------------------------------------------------------------------
# Canned Responses (jawaban template)
# ---------------------------------------------------------------------------

@login_required
def canned_response_list(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    profile = request.user.profile
    if request.user.is_superuser:
        responses = CannedResponse.objects.select_related('category', 'company').all()
    else:
        responses = CannedResponse.objects.select_related('category', 'company').filter(
            Q(company=profile.company) | Q(company__isnull=True)
        )

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        category_id = request.POST.get('category', '')
        company_id = request.POST.get('company', '')
        if not title or not content:
            messages.error(request, 'Judul dan isi wajib diisi.')
        else:
            CannedResponse.objects.create(
                title=title,
                content=content,
                category_id=category_id or None,
                company_id=company_id or None,
                created_by=request.user,
            )
            messages.success(request, 'Template jawaban berhasil dibuat.')
            return redirect('canned_response_list')

    return render(request, 'tickets/canned_response_list.html', {
        'responses': responses,
        'categories': Category.objects.all(),
        'companies': Company.objects.all(),
    })


@login_required
def canned_response_edit(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    response = get_object_or_404(CannedResponse, pk=pk)
    if request.method == 'POST':
        response.title = request.POST.get('title', response.title).strip()
        response.content = request.POST.get('content', response.content).strip()
        response.category_id = request.POST.get('category') or None
        response.company_id = request.POST.get('company') or None
        response.save()
        messages.success(request, 'Template berhasil diupdate.')
        return redirect('canned_response_list')

    return render(request, 'tickets/canned_response_form.html', {
        'response': response,
        'categories': Category.objects.all(),
        'companies': Company.objects.all(),
    })


@login_required
def canned_response_delete(request, pk):
    if not admin_required(request.user):
        return redirect('dashboard')

    if request.method == 'POST':
        CannedResponse.objects.filter(pk=pk).delete()
        messages.success(request, 'Template berhasil dihapus.')
    return redirect('canned_response_list')


# ---------------------------------------------------------------------------
# Print Ticket PDF
# ---------------------------------------------------------------------------

@login_required
def ticket_print_pdf(request, pk):
    ticket = get_visible_ticket(request, pk)

    response = HttpResponse(content_type='application/pdf')
    filename = f"tiket-{ticket.id}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(f"Tiket #{ticket.id}", styles['Title']))
    elements.append(Paragraph(f"{ticket.title}", styles['Heading2']))
    elements.append(Paragraph(" ", styles['Normal']))

    info = [
        ['Company', ticket.company.name],
        ['Kategori', ticket.category.name if ticket.category else '-'],
        ['Status', ticket.get_status_display()],
        ['Prioritas', ticket.get_priority_display()],
        ['Dibuat Oleh', ticket.created_by.username],
        ['Assigned To', ticket.assigned_to.username if ticket.assigned_to else 'Belum ada'],
        ['Dibuat Pada', ticket.created_at.strftime('%d %B %Y %H:%M')],
        ['SLA Deadline', ticket.sla_deadline.strftime('%d %B %Y %H:%M') if ticket.sla_deadline else '-'],
    ]
    if ticket.closed_at:
        info.append(['Ditutup', ticket.closed_at.strftime('%d %B %Y %H:%M')])

    table = Table(info, colWidths=[4*cm, 12*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
    ]))
    elements.append(table)
    elements.append(Paragraph(" ", styles['Normal']))

    elements.append(Paragraph("Deskripsi", styles['Heading3']))
    for line in ticket.description.split('\n'):
        elements.append(Paragraph(line or '&nbsp;', styles['Normal']))
    elements.append(Paragraph(" ", styles['Normal']))

    comments = ticket.comments.select_related('author').order_by('created_at')
    if comments:
        elements.append(Paragraph(f"Komentar ({comments.count()})", styles['Heading3']))
        for c in comments:
            elements.append(Paragraph(
                f"<b>{c.author.username}</b> ({c.created_at.strftime('%d/%m/%Y %H:%M')}): {c.message[:300]}",
                styles['Normal']
            ))
            elements.append(Paragraph(" ", styles['Normal']))

    time_entries = ticket.time_entries.select_related('user').order_by('-started_at')
    if time_entries:
        elements.append(Paragraph("Waktu Kerja", styles['Heading3']))
        time_data = [['Staff', 'Deskripsi', 'Durasi', 'Tanggal']]
        for te in time_entries:
            time_data.append([
                te.user.username,
                te.description or '-',
                f"{te.duration_minutes} menit",
                te.started_at.strftime('%d/%m/%Y %H:%M'),
            ])
        time_table = Table(time_data, colWidths=[3*cm, 6*cm, 3*cm, 4*cm])
        time_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(time_table)

    doc.build(elements)
    return response


# ---------------------------------------------------------------------------
# Auto-Assignment Rules
# ---------------------------------------------------------------------------

@login_required
def auto_assign_rules(request):
    if not superuser_required(request.user):
        return redirect('dashboard')

    rules = AutoAssignRule.objects.select_related('company', 'category', 'assign_to').all()

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        company_id = request.POST.get('company', '')
        category_id = request.POST.get('category', '')
        priority = request.POST.get('priority', '')
        assign_to_id = request.POST.get('assign_to', '')
        if not name or not company_id or not assign_to_id:
            messages.error(request, 'Nama, company, dan staff wajib diisi.')
        else:
            AutoAssignRule.objects.create(
                name=name,
                company_id=company_id,
                category_id=category_id or None,
                priority=priority or '',
                assign_to_id=assign_to_id,
            )
            messages.success(request, 'Aturan auto-assign berhasil dibuat.')
            return redirect('auto_assign_rules')

    return render(request, 'tickets/auto_assign_rules.html', {
        'rules': rules,
        'companies': Company.objects.all(),
        'categories': Category.objects.all(),
        'staff_users': User.objects.filter(profile__role__in=['staff', 'admin']),
    })


@login_required
def auto_assign_rule_delete(request, pk):
    if not superuser_required(request.user):
        return redirect('dashboard')

    if request.method == 'POST':
        AutoAssignRule.objects.filter(pk=pk).delete()
        messages.success(request, 'Aturan berhasil dihapus.')
    return redirect('auto_assign_rules')


# ---------------------------------------------------------------------------
# Real-time Notifications (AJAX polling)
# ---------------------------------------------------------------------------

@login_required
def api_notifications(request):
    """Return unread notifications + count for AJAX polling."""
    notifs = request.user.notifications.filter(is_read=False).order_by('-created_at')[:10]
    return JsonResponse({
        'count': request.user.notifications.filter(is_read=False).count(),
        'notifications': [
            {
                'id': n.id,
                'message': n.message,
                'ticket_id': n.ticket_id,
                'created_at': n.created_at.strftime('%d %b %Y %H:%M'),
            }
            for n in notifs
        ],
    })


@login_required
def api_mark_notifications_read(request):
    """Mark all notifications as read."""
    if request.method == 'POST':
        request.user.notifications.filter(is_read=False).update(is_read=True)
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=405)


# ---------------------------------------------------------------------------
# Audit Log Export
# ---------------------------------------------------------------------------

@login_required
def activity_export_excel(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    activities = Activity.objects.select_related('ticket', 'user', 'ticket__company').order_by('-created_at')
    if not request.user.is_superuser:
        activities = activities.filter(ticket__company=request.user.profile.company)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Audit Log'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    headers = ['Waktu', 'Tiket', 'User', 'Aksi', 'Detail', 'Company']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row, act in enumerate(activities[:1000], 2):
        ws.cell(row=row, column=1, value=act.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row, column=2, value=f'#{act.ticket.id}')
        ws.cell(row=row, column=3, value=act.user.username if act.user else 'Sistem')
        ws.cell(row=row, column=4, value=act.get_action_display())
        ws.cell(row=row, column=5, value=act.detail)
        ws.cell(row=row, column=6, value=act.ticket.company.name if act.ticket.company else '')

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="audit_log.xlsx"'
    wb.save(response)
    return response


@login_required
def activity_export_csv(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    activities = Activity.objects.select_related('ticket', 'user', 'ticket__company').order_by('-created_at')
    if not request.user.is_superuser:
        activities = activities.filter(ticket__company=request.user.profile.company)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="audit_log.csv"'
    writer = csv.writer(response)
    writer.writerow(['Waktu', 'Tiket', 'User', 'Aksi', 'Detail', 'Company'])
    for act in activities[:1000]:
        writer.writerow([
            act.created_at.strftime('%d/%m/%Y %H:%M'),
            f'#{act.ticket.id}',
            act.user.username if act.user else 'Sistem',
            act.get_action_display(),
            act.detail,
            act.ticket.company.name if act.ticket.company else '',
        ])
    return response


@login_required
def activity_export_pdf(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    activities = Activity.objects.select_related('ticket', 'user', 'ticket__company').order_by('-created_at')
    if not request.user.is_superuser:
        activities = activities.filter(ticket__company=request.user.profile.company)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="audit_log.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph('Audit Log', styles['Title'])]

    data = [['Waktu', 'Tiket', 'User', 'Aksi', 'Detail', 'Company']]
    for act in activities[:500]:
        data.append([
            act.created_at.strftime('%d/%m/%Y %H:%M'),
            f'#{act.ticket.id}',
            act.user.username if act.user else 'Sistem',
            act.get_action_display(),
            act.detail[:60],
            act.ticket.company.name if act.ticket.company else '',
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(table)
    doc.build(elements)
    return response


# ---------------------------------------------------------------------------
# SLA Escalation Management
# ---------------------------------------------------------------------------

@login_required
def sla_escalation_log(request):
    if not admin_required(request.user):
        return redirect('dashboard')

    escalated = Ticket.objects.filter(
        sla_overdue_sent=True,
        status__in=['open', 'in_progress'],
    ).select_related('company', 'category', 'assigned_to', 'created_by').order_by('-created_at')

    if not request.user.is_superuser:
        escalated = escalated.filter(company=request.user.profile.company)

    paginator = Paginator(escalated, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'tickets/sla_escalation_log.html', {
        'tickets': page_obj,
    })
